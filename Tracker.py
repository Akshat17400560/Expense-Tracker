import json
import os
import datetime
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox
from tkcalendar import DateEntry

JSON_FILE = "expenses.json"
EXCEL_FILE = "expenses.xlsx"

#Data Handling 
def load_expenses():
    if not os.path.exists(JSON_FILE):
        return []
    with open(JSON_FILE, 'r') as f:
        return json.load(f)

def save_expenses(expenses):
    with open(JSON_FILE, 'w') as f:
        json.dump(expenses, f, indent=4)

def write_to_excel(expense):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount", "Category", "Date"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    ws.append([expense["amount"], expense["category"], expense["date"]])
    wb.save(EXCEL_FILE)

# GUI Functions 
def add_expense_gui():
    try:
        amount_str = entry_amount.get()
        category = entry_category.get()
        date_val = date_picker.get_date()

        if not amount_str or not category:
            messagebox.showwarning("Warning", "Please fill all fields.")
            return

        amount = float(amount_str)

        expense = {
            "amount": amount,
            "category": category,
            "date": str(date_val)
        }

        expenses.append(expense)
        save_expenses(expenses)
        write_to_excel(expense)

        messagebox.showinfo("Success", "✅ Expense added!")
        entry_amount.delete(0, tk.END)
        entry_category.delete(0, tk.END)

    except ValueError:
        messagebox.showerror("Error", "Please enter a valid number for amount.")
    except Exception as e:
        messagebox.showerror("Error", f"❌ {e}")

def view_expenses_gui():
    if not expenses:
        messagebox.showinfo("Expenses", "No expenses recorded yet.")
        return
    result = "\n".join([f"₹{exp['amount']} | {exp['category']} | {exp['date']}" for exp in expenses])
    messagebox.showinfo("📋 All Expenses", result)

def show_summary_gui():
    def on_submit():
        try:
            month = int(month_entry.get())
            year = int(year_entry.get())
            if not (1 <= month <= 12):
                raise ValueError("Month must be between 1 and 12.")
            summary_win.destroy()
            show_summary_for_month_year(month, year)
        except Exception as e:
            messagebox.showerror("Input Error", str(e))

    summary_win = tk.Toplevel(app)
    summary_win.title("Select Month and Year")
    summary_win.geometry("250x120")
    summary_win.grab_set()

    tk.Label(summary_win, text="Month (1-12):").grid(row=0, column=0, padx=10, pady=10, sticky='e')
    month_entry = tk.Entry(summary_win)
    month_entry.grid(row=0, column=1)

    tk.Label(summary_win, text="Year (e.g., 2025):").grid(row=1, column=0, padx=10, pady=5, sticky='e')
    year_entry = tk.Entry(summary_win)
    year_entry.grid(row=1, column=1)

    submit_btn = tk.Button(summary_win, text="Show Summary", command=on_submit, bg="#17a2b8", fg="white")
    submit_btn.grid(row=2, column=0, columnspan=2, pady=10)

    summary_win.bind('<Return>', lambda event: on_submit())

def show_summary_for_month_year(month, year):
    try:
        filtered = [
            exp for exp in expenses
            if datetime.datetime.strptime(exp['date'], "%Y-%m-%d").month == month
            and datetime.datetime.strptime(exp['date'], "%Y-%m-%d").year == year
        ]

        if not filtered:
            messagebox.showinfo("Summary", "No expenses found for the given month.")
            return

        totals = defaultdict(float)
        for exp in filtered:
            totals[exp['category']] += exp['amount']

        summary = "\n".join([f"{cat}: ₹{amt:.2f}" for cat, amt in totals.items()])
        messagebox.showinfo(f"📊 Summary for {month}/{year}", summary)

        # Pie chart
        plt.figure(figsize=(6, 6))
        plt.pie(totals.values(), labels=totals.keys(), autopct='%1.1f%%', startangle=140)
        plt.title(f"Expenses Breakdown for {month}/{year}")
        plt.axis('equal')
        plt.show()

    except Exception as e:
        messagebox.showerror("Error", f"❌ {e}")

# GUI Setup
expenses = load_expenses()

app = tk.Tk()
app.title("💰 Expense Tracker")
app.geometry("350x300")
app.config(padx=10, pady=10)
app.bell = lambda *args, **kwargs: None  # Disable system beep on Enter

# Form Fields
tk.Label(app, text="Amount (₹):").grid(row=0, column=0, sticky="w", pady=(5, 0))
entry_amount = tk.Entry(app)
entry_amount.grid(row=0, column=1, pady=(5, 0))

tk.Label(app, text="Category:").grid(row=1, column=0, sticky="w")
entry_category = tk.Entry(app)
entry_category.grid(row=1, column=1)

tk.Label(app, text="Date:").grid(row=2, column=0, sticky="w")
date_picker = DateEntry(app, width=17, background="darkblue", foreground="white", date_pattern="yyyy-mm-dd")
date_picker.grid(row=2, column=1)

# Buttons
tk.Button(app, text="Add Expense", command=add_expense_gui, bg="#28a745", fg="white").grid(row=3, column=0, columnspan=2, pady=10, sticky="ew")
tk.Button(app, text="View Expenses", command=view_expenses_gui, bg="#007bff", fg="white").grid(row=4, column=0, columnspan=2, pady=5, sticky="ew")
tk.Button(app, text="Monthly Summary with Chart", command=show_summary_gui, bg="#ffc107").grid(row=5, column=0, columnspan=2, pady=5, sticky="ew")
tk.Button(app, text="Exit", command=app.quit, bg="#dc3545", fg="white").grid(row=6, column=0, columnspan=2, pady=10, sticky="ew")

app.mainloop()
